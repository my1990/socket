import requests
import openpyxl
import os

if os.path.exists('阿里job.xlsx') == False:
    excel = openpyxl.Workbook()
    sheet = excel.active
    sheet.title = '技术类工作'
    excel.save(filename='阿里job.xlsx')

def login():
    url = 'https://job.alibaba.com/zhaopin/socialPositionList/doList.json'
    post_data = {
        'pageSize':'10',
        'pageIndex':1,
        'location':'杭州',
        'first':'技术类'
    }
    t = requests.post(url, data=post_data).json()
    totalPage = t['returnValue']['totalPage']
    print('总页数:',totalPage)
    list = ['系统运维','运维','linux','Linux']
    page = 1
    while page <= totalPage:
    # if page == 1:
        post_data = {
            'pageSize': '10',
            'pageIndex': page,
            'location': '杭州',
            'first': '技术类'
        }
        h = requests.post(url, data=post_data, timeout=30).json()
        text_list = h['returnValue']['datas']
        for job in text_list:
            job_name = job['name']                  #职位名称
            job_city = job['workLocation']         #工作地点
            job_description = job['description'].replace('<br/>','')   #工作描述
            job_requirement = job['requirement'].replace('<br/>','')    #职位要求
            for k in list:
                if k in job_name:  #如果职位名称里存在list列表中关键字,则保存相关信息
                    write_excel(job_name,job_city,job_description,job_requirement)

        print('保存第%s页信息,共%s页,还剩%s页~' % (page,totalPage,totalPage-page))
        page += 1


#a为职位名称  b为工作地点 c为工作描述  d为职位要求
def write_excel(a,b,c,d):
    excel1 = openpyxl.load_workbook('阿里job.xlsx')
    sheet1 = excel1.active
    row_num = sheet1.max_row + 2 #查看已经有多少行数据.从下下一行开始写

    sheet1.cell(row=row_num, column=1, value='职位名称')  #row行  cou 列
    sheet1.cell(row=row_num, column=2, value=a)
    sheet1.cell(row=row_num + 1, column=1, value='工作地点')
    sheet1.cell(row=row_num + 1, column=2, value=b)
    sheet1.cell(row=row_num + 2, column=1, value='工作描述')
    sheet1.cell(row=row_num + 2, column=2, value=c)
    sheet1.cell(row=row_num + 3, column=1, value='职位要求')
    sheet1.cell(row=row_num + 3, column=2, value=d)

    excel1.save(filename='阿里job.xlsx')

if __name__ == "__main__":
    login()


